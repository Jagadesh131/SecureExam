import pytest
import time
import os
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Global list to store test results
test_results = []

@pytest.fixture(scope="session")
def browser_config():
    """Setup Chrome Options once per session"""
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--disable-software-rasterizer")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--window-size=1920,1080")
    return chrome_options

@pytest.fixture()
def browser(browser_config):
    """Provide a fresh WebDriver instance per test"""
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=browser_config)
    driver.implicitly_wait(5)
    
    yield driver
    
    driver.quit()

@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    """Intercept test results to build the Excel report"""
    outcome = yield
    report = outcome.get_result()
    
    # We only care about the actual test execution phase (not setup/teardown)
    # or setup failures.
    if report.when == "call" or (report.when == "setup" and report.failed):
        # Determine Module from filename
        module_name = os.path.basename(item.location[0])
        
        # Determine Test ID and Description
        test_id = item.name
        description = item.function.__doc__ if item.function.__doc__ else item.name
        
        status = "Pass" if report.passed else "Fail"
        
        # Format the error message if failed
        error_msg = ""
        if report.failed:
            error_msg = str(report.longreprtext)
            # Truncate very long errors for Excel readability
            if len(error_msg) > 500:
                error_msg = error_msg[:500] + "... [truncated]"

        # Determine Test Category
        test_id_upper = test_id.upper()
        if "UNIT_" in test_id_upper or "TEST_UNIT" in item.location[0].upper():
            category = "Unit Testing"
        elif "AUTH_" in test_id_upper or "SQLI" in test_id_upper or "XSS" in test_id_upper:
            category = "Validation Test"
        elif "UI" in test_id_upper or "DASHBOARD" in test_id_upper or "RENDER" in test_id_upper:
            category = "UI/UX Test"
        else:
            category = "Functional Testing"

        test_results.append({
            "Test ID": test_id,
            "Module": module_name,
            "Test Category": category,
            "Test Case Description": description.strip(),
            "Status": status,
            "Execution Time (s)": round(report.duration, 2),
            "Error Message": error_msg
        })

def pytest_sessionfinish(session, exitstatus):
    """Export results to Excel at the end of the test session with advanced styling"""
    if test_results:
        df = pd.DataFrame(test_results)
        
        # Calculate Summary Metrics
        total_tests = len(df)
        passed_tests = len(df[df['Status'] == 'Pass'])
        failed_tests = total_tests - passed_tests
        pass_rate = (passed_tests / total_tests) * 100 if total_tests > 0 else 0
        pass_rate_str = f"{pass_rate:.1f}%"
        
        duration = df['Execution Time (s)'].sum()
        duration_str = f"{duration:.2f}"
        
        # Save to the root of the backend directory
        report_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', 'E2E_Test_Report.xlsx'))
        
        try:
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                # We let pandas write the detailed results first
                df.to_excel(writer, sheet_name='Detailed Results', index=False)
                
                # Format Detailed Results sheet
                detailed_ws = writer.sheets['Detailed Results']
                detailed_ws.column_dimensions['A'].width = 30
                detailed_ws.column_dimensions['B'].width = 30
                detailed_ws.column_dimensions['C'].width = 20
                detailed_ws.column_dimensions['D'].width = 50
                detailed_ws.column_dimensions['E'].width = 15
                detailed_ws.column_dimensions['F'].width = 15
                detailed_ws.column_dimensions['G'].width = 80
                
                # Create the Summary Sheet manually
                wb = writer.book
                summary_ws = wb.create_sheet('Summary', 0)
                
                # --- Styling Definitions ---
                dark_blue_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
                light_grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                
                white_bold = Font(color="FFFFFF", bold=True, size=16)
                header_font = Font(color="FFFFFF", bold=True, size=11)
                title_font = Font(color="1F497D", bold=True, size=14)
                blue_bold_large = Font(color="1F497D", bold=True, size=24)
                green_bold_large = Font(color="385723", bold=True, size=24)
                red_bold_large = Font(color="C00000", bold=True, size=24)
                
                green_bold_status = Font(color="385723", bold=True, size=11)
                red_bold_status = Font(color="C00000", bold=True, size=11)
                
                center_align = Alignment(horizontal="center", vertical="center")
                left_align = Alignment(horizontal="left", vertical="center")
                
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                # --- Adjust Column Widths ---
                summary_ws.column_dimensions['A'].width = 25
                summary_ws.column_dimensions['B'].width = 30
                summary_ws.column_dimensions['C'].width = 20
                summary_ws.column_dimensions['D'].width = 20
                summary_ws.column_dimensions['E'].width = 20
                
                # --- 1. Main Header ---
                summary_ws.merge_cells('A1:E1')
                cell = summary_ws['A1']
                cell.value = "SECUREEXAM APP - QA E2E TEST REPORT"
                cell.fill = dark_blue_fill
                cell.font = white_bold
                cell.alignment = center_align
                
                # --- 2. Metadata Block ---
                summary_ws['A3'] = "Test Run Date:"
                summary_ws['B3'] = time.strftime("%Y-%m-%d")
                summary_ws['A4'] = "Test Run Time:"
                summary_ws['B4'] = time.strftime("%H:%M:%S")
                summary_ws['A5'] = "OS / Platform:"
                summary_ws['B5'] = "Windows / Python Test Server"
                summary_ws['A6'] = "App Version Name:"
                summary_ws['B6'] = "1.0 (Universal)"
                
                for r in range(3, 7):
                    summary_ws[f'A{r}'].font = Font(bold=True)
                    summary_ws[f'B{r}'].alignment = left_align
                    
                summary_ws['A7'] = "Deployable Status:"
                summary_ws['A7'].font = Font(bold=True)
                
                status_cell = summary_ws['B7']
                summary_ws.merge_cells('B7:E7')
                if failed_tests == 0:
                    status_cell.value = "DEPLOYABLE - FIT FOR RELEASE"
                    status_cell.fill = green_fill
                    status_cell.font = green_bold_status
                else:
                    status_cell.value = "FAILED - DO NOT DEPLOY"
                    status_cell.fill = red_fill
                    status_cell.font = red_bold_status
                    
                # --- 3. Core Metrics KPI Summary ---
                summary_ws['A10'] = "Core Metrics KPI Summary"
                summary_ws['A10'].font = title_font
                
                headers_kpi = ["TOTAL TEST CASES", "PASSED", "FAILED", "PASS RATE", "DURATION (SEC)"]
                for col_idx, header in enumerate(headers_kpi, 1):
                    cell = summary_ws.cell(row=11, column=col_idx)
                    cell.value = header
                    cell.fill = light_grey_fill
                    cell.font = Font(color="1F497D", bold=True, size=10)
                    cell.alignment = center_align
                    cell.border = thin_border
                    
                values_kpi = [
                    (total_tests, blue_bold_large),
                    (passed_tests, green_bold_large),
                    (failed_tests, red_bold_large if failed_tests > 0 else blue_bold_large),
                    (pass_rate_str, blue_bold_large),
                    (duration_str, blue_bold_large)
                ]
                
                for col_idx, (val, font) in enumerate(values_kpi, 1):
                    cell = summary_ws.cell(row=12, column=col_idx)
                    cell.value = val
                    cell.font = font
                    cell.alignment = center_align
                    cell.border = thin_border
                    
                summary_ws.row_dimensions[12].height = 40
                
                # --- 4. Category-Wise Execution Breakdown ---
                summary_ws['A15'] = "Category-Wise Execution Breakdown"
                summary_ws['A15'].font = title_font
                
                headers_cat = ["Test Category", "Total Cases", "Passed", "Failed", "Pass Rate"]
                for col_idx, header in enumerate(headers_cat, 1):
                    cell = summary_ws.cell(row=16, column=col_idx)
                    cell.value = header
                    cell.fill = dark_blue_fill
                    cell.font = header_font
                    cell.alignment = center_align if col_idx > 1 else left_align
                    cell.border = thin_border
                
                categories = ["Validation Test", "UI/UX Test", "Functional Testing", "Unit Testing"]
                current_row = 17
                for cat in categories:
                    cat_df = df[df['Test Category'] == cat]
                    t = len(cat_df)
                    if t == 0: continue
                    p = len(cat_df[cat_df['Status'] == 'Pass'])
                    f = t - p
                    rate = f"{(p/t)*100:.1f}%"
                    
                    row_data = [cat, t, p, f, rate]
                    for col_idx, val in enumerate(row_data, 1):
                        cell = summary_ws.cell(row=current_row, column=col_idx)
                        cell.value = val
                        cell.alignment = center_align if col_idx > 1 else left_align
                        cell.border = thin_border
                    current_row += 1
                    
                # Total Summary Row
                row_data = ["Total Summary", total_tests, passed_tests, failed_tests, pass_rate_str]
                for col_idx, val in enumerate(row_data, 1):
                    cell = summary_ws.cell(row=current_row, column=col_idx)
                    cell.value = val
                    cell.font = Font(bold=True)
                    cell.alignment = center_align if col_idx > 1 else left_align
                    cell.border = thin_border
                
            print(f"\nSuccessfully generated styled E2E Test Report at: {report_path}")
        except Exception as e:
            print(f"\nFailed to generate E2E Excel Report: {e}")
